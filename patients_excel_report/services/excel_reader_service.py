"""
Port of ExcelReaderService.cs

The main Excel processing engine used by PatientReaderReportController.
For every sheet in the workbook it:
  1. Locates the header row (Sun/Mon/Sat labels)
  2. Finds the "Patient Signature" and "Service Provider Signature" rows
  3. Extracts service_provider_name, patient, contract, caregiver-id metadata
  4. For each of the 7 day columns (F, H, K, N, O, T, W):
     - reads date, personal-care code(s), mobile text note
     - detects copy-paste in mobile notes via text similarity
     - extracts embedded images at signature cell addresses
     - pixel-analyses images to determine signed/unsigned
     - validates codes + notes via CareValidationService
  5. Returns a list of ReportRow dicts (one per day per sheet)
"""

import logging
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl import load_workbook

from django.conf import settings

from .care_validation_service import CareValidationService
from .signature_detection_service import SignatureDetectionService
from .text_similarity_service import TextSimilarityService
from .image_extractor import extract_images_anchored_to_worksheet

logger = logging.getLogger("patients_excel_report.excel_reader_service")

# ---------------------------------------------------------------------------
# Day-column map  (0-based column indices for F,H,K,N,O,T,W)
# ---------------------------------------------------------------------------
DAY_MAP: list[tuple[int, str]] = [
    (5,  "Sun"),
    (7,  "Mon"),
    (10, "Tue"),
    (13, "Wed"),
    (14, "Thu"),
    (19, "Fri"),
    (22, "Sat"),
]

CONTRACTS_WITH_CAREGIVER_SIG = {
    "MCCMH CW", "MCCMH SCFS", "MCCMH SED",
    "GT Independence", "GT Independence Respite",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _normalize_whitespace(s: str) -> str:
    if not s or not s.strip():
        return ""
    return " ".join(s.split()).strip()


def _column_letter_from_index(zero_based_index: int) -> str:
    index = zero_based_index + 1
    col = ""
    while index > 0:
        rem = (index - 1) % 26
        col = chr(ord("A") + rem) + col
        index = (index - 1) // 26
    return col


def _safe_str(val) -> str:
    return str(val).strip() if val is not None else ""

def _format_time(raw: str) -> str:
    if not raw:
        return raw

    # openpyxl returns Excel time fractions as floats (0.0 - 1.0)
    try:
        fraction = float(raw)
        if 0.0 <= fraction < 1.0:
            total_minutes = round(fraction * 24 * 60)
            hh = total_minutes // 60
            mm = total_minutes % 60
            return f"{hh:02d}:{mm:02d}"
    except (ValueError, TypeError):
        pass

    # Strip non-digit chars to normalise "09:00", "9.00", "0900"
    digits = "".join(c for c in raw if c.isdigit())

    if len(digits) == 3:      # "900"  → "09:00"
        digits = digits.zfill(4)
    if len(digits) == 4:      # "0900" → "09:00"
        return f"{digits[:2]}:{digits[2:]}"

    return raw  # already formatted or unrecognised


class ExcelReaderService:
    """
    Processes an .xlsx file and returns a list of ReportRow dicts.
    """

    def __init__(
        self,
        care_service: Optional[CareValidationService] = None,
        sig_service: Optional[SignatureDetectionService] = None,
        text_sim_service: Optional[TextSimilarityService] = None,
    ):
        cfg = getattr(settings, "EXCEL_PROCESSING", {})
        self._max_workers: int = cfg.get("MAX_DEGREE_OF_PARALLELISM", 4)
        self._threshold: float = cfg.get("SIMILARITY_THRESHOLD", 0.75)

        self._care = care_service or CareValidationService()
        self._sig = sig_service or SignatureDetectionService()
        self._text_sim = text_sim_service or TextSimilarityService()

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def process_file(self, file_path: str) -> list[dict]:
        """
        Parse *file_path* (.xlsx) and return a list of row dicts.
        Returns [] on failure.
        """
        try:
            # ExcelDataReader in .NET reads raw values; openpyxl with
            # data_only=True gives us evaluated cell values.
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            logger.error("Failed to open workbook %s: %s", file_path, exc)
            return []

        all_rows: list[dict] = []

        for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
            try:
                ws = wb[sheet_name]
                sheet_rows = self._process_sheet(file_path, ws, sheet_name)
                all_rows.extend(sheet_rows)
            except Exception as exc:
                logger.error("Error processing sheet %s: %s", sheet_name, exc)

        wb.close()
        return all_rows

    # ------------------------------------------------------------------
    # Sheet processing
    # ------------------------------------------------------------------
    def _process_sheet(self, file_path: str, ws, sheet_name: str) -> list[dict]:
        # Load all cells into a buffer (list of rows, each row is a list of values)
        sheet_data: list[list] = []
        for row in ws.iter_rows(values_only=True):
            sheet_data.append(list(row))

        if not sheet_data:
            return []

        # ---- "No duties performed" flag ----
        duties_not_performed = any(
            _safe_str(cell).lower() == "no duties performed."
            for row in sheet_data
            for cell in row
        )

        # ---- Find header row (contains Sun, Mon, Sat) ----
        header_row = -1
        for r_idx, row in enumerate(sheet_data):
            str_vals = [_safe_str(c) for c in row]
            if "Sun" in str_vals and "Mon" in str_vals and "Sat" in str_vals:
                header_row = r_idx
                break
        if header_row == -1:
            return []

        date_row = header_row + 2

        # ---- Find "Patient Signature" row ----
        sig_row = -1
        for r_idx, row in enumerate(sheet_data):
            combined = _normalize_whitespace(
                " ".join(_safe_str(row[c]) for c in range(min(5, len(row))))
            )
            if combined.lower() == "patient signature":
                sig_row = r_idx
                break
        if sig_row == -1:
            return []

        # ---- Extract metadata ----
        caregiver_id = self._extract_value_from_buffer(sheet_data, "Code", search_next_row=True)
        service_provider_name = self._extract_value_from_buffer(sheet_data, "Service Provider Name", search_next_row=True)
        patient = self._extract_value_from_buffer(sheet_data, "Patient Name:", search_next_row=False)
        contract = self._extract_contract(sheet_data)

        # ---- Find "Service Provider Signature" row (only for certain contracts) ----
        caregiver_sig_row = -1
        for r_idx, row in enumerate(sheet_data):
            combined = _normalize_whitespace(
                " ".join(_safe_str(row[c]) for c in range(min(5, len(row))))
            )
            if combined.lower() == "service provider signature":
                caregiver_sig_row = r_idx
                break
        if caregiver_sig_row == -1:
            return []

        # ---- Personal Care rows ----
        personal_care_row_indices: list[int] = []
        mobile_text_row_index = -1

        for r_idx, row in enumerate(sheet_data):
            label = _safe_str(row[0]) if row else ""
            if label.lower() == "personal care":
                personal_care_row_indices.append(r_idx)
            if mobile_text_row_index == -1 and label.lower() == "mobile text note":
                mobile_text_row_index = r_idx

        is_personal_care_column_present = len(personal_care_row_indices) > 0

        # ---- Respite T1005 rows ----
        pc_respite_rows: list[int] = []
        for pc_row_idx in personal_care_row_indices:
            row = sheet_data[pc_row_idx]
            for c_idx in range(1, len(row)):
                val = _safe_str(row[c_idx])
                if val.lower() == "respite t1005":
                    pc_respite_rows.append(pc_row_idx)

        # ---- Start/End Time rows ----
        start_time_row_index = self._find_row_by_leading_label(sheet_data, "Start Time")
        end_time_row_index = self._find_row_by_leading_label(sheet_data, "End Time")

        # ---- Extract images ----
        try:
            images = extract_images_anchored_to_worksheet(file_path, sheet_name)
        except Exception as exc:
            logger.warning("Image extraction failed for sheet %s: %s", sheet_name, exc)
            images = {}

        # ---- Process each day column (parallel) ----
        results: list[dict] = []
        day_indices = list(range(len(DAY_MAP)))

        def process_day(di: int) -> Optional[dict]:
            try:
                col_index, day_key = DAY_MAP[di]

                # Date
                date_val = ""
                if date_row < len(sheet_data) and col_index < len(sheet_data[date_row]):
                    date_val = _safe_str(sheet_data[date_row][col_index])

                # Personal Care values (multi-row)
                pc_values: list[str] = []
                pc_descs_and_values: list[tuple[int, int, str, str]] = []
                for pc_row_idx in personal_care_row_indices:
                    row = sheet_data[pc_row_idx]
                    if col_index < len(row):
                        val = str(row[col_index]).strip() if row[col_index] is not None else ""
                        pc_values.append(val)
                        desc = _safe_str(row[2]) if len(row) > 2 else ""
                        pc_descs_and_values.append((pc_row_idx, col_index, desc, val))

                # Mobile Text Note
                mobile = ""
                if mobile_text_row_index >= 0:
                    mobile_row = sheet_data[mobile_text_row_index]
                    if col_index < len(mobile_row):
                        mobile = _safe_str(mobile_row[col_index])

                # Copy-paste detection
                is_likely_copy_paste = False
                if mobile and mobile_text_row_index >= 0:
                    mobile_row = sheet_data[mobile_text_row_index]
                    for c_idx in range(1, len(mobile_row)):
                        if c_idx == col_index:
                            continue
                        other = _safe_str(mobile_row[c_idx])
                        if not other:
                            continue
                        if self._text_sim.is_above_threshold(mobile, other, self._threshold):
                            is_likely_copy_paste = True
                            break

                # Service type
                service_type_cls = False
                service_type_respite = False
                for (_, pc_col, desc, _) in pc_descs_and_values:
                    if pc_col == col_index:
                        if "cls" in desc.lower():
                            service_type_cls = True
                        if "respite" in desc.lower():
                            service_type_respite = True

                if service_type_cls and service_type_respite:
                    type_of_service = "cls, respite"
                elif service_type_cls:
                    type_of_service = "cls"
                elif service_type_respite:
                    type_of_service = "respite"
                else:
                    type_of_service = ""

                # Respite T1005 check for mobile validation
                only_respite_t1005 = [
                    x for x in pc_descs_and_values if x[2] == "Respite T1005"
                ]
                check_respite_for_mobile = [
                    x for x in only_respite_t1005
                    if x[1] == col_index and len(mobile) > 0
                ]

                # Validation
                pc_status, mobile_status, final_result = self._care.validate_multiple(
                    pc_values,
                    mobile,
                    duties_not_performed,
                    is_personal_care_column_present,
                    check_respite_for_mobile,
                )

                # Signature detection – patient
                cell_addr = _column_letter_from_index(col_index) + str(sig_row + 1)
                img_bytes = images.get(cell_addr)
                signature_status = (
                    self._sig.analyze_image_non_empty_pixel_percent(img_bytes)
                    if img_bytes
                    else "unsigned"
                )

                # Signature detection – service_provider_name
                cg_cell_addr = (
                    _column_letter_from_index(col_index) + str(caregiver_sig_row + 1)
                    if caregiver_sig_row >= 0
                    else ""
                )
                cg_img_bytes = images.get(cg_cell_addr) if cg_cell_addr else None
                caregiver_signature_status = (
                    self._sig.analyze_image_non_empty_pixel_percent(cg_img_bytes)
                    if cg_img_bytes
                    else "unsigned"
                )

                # Start / End Time
                start_time = ""
                start_time_gps = ""
                if start_time_row_index >= 0:
                    st_row = sheet_data[start_time_row_index]
                    if col_index < len(st_row):
                        st_label = _safe_str(st_row[0])
                        st_val = _safe_str(st_row[col_index])
                        if st_label == "Start Time":
                            start_time = _format_time(st_val)
                        elif st_label == "Start Time (GPS Coordinates)":
                            if len(st_val) >= 4:
                                start_time = st_val[:4]
                                start_time_gps = st_val[4:].strip() if len(st_val) > 4 else ""

                end_time = ""
                end_time_gps = ""
                if end_time_row_index >= 0:
                    et_row = sheet_data[end_time_row_index]
                    if col_index < len(et_row):
                        et_label = _safe_str(et_row[0])
                        et_val = _safe_str(et_row[col_index])
                        if et_label == "End Time":
                            end_time = _format_time(et_val)
                        elif et_label == "End Time (GPS Coordinates)":
                            if len(et_val) >= 4:
                                end_time = et_val[:4]
                                end_time_gps = et_val[4:].strip() if len(et_val) > 4 else ""

                # Highlight flag
                highlight = (
                    bool(start_time) and bool(end_time)
                    and (pc_status.lower() == "no value inserted" or duties_not_performed)
                    and mobile_status.lower() == "text absent"
                    and signature_status.lower() == "unsigned"
                )

                return {
                    "sheetName": sheet_name,
                    "service_provider_name": service_provider_name,
                    "patient": patient,
                    "contracts": contract,
                    "day": day_key,
                    "date": date_val,
                    "personalCareStatus": pc_status,
                    "mobileNoteValue": mobile_status,
                    "signatureStatus": signature_status,
                    "caregiverSignatureStatus": caregiver_signature_status,
                    "validationResult": final_result,
                    "highlightColumn": highlight,
                    "startTime": start_time,
                    "endTime": end_time,
                    "startTimeGPSCoordinates": start_time_gps,
                    "endTimeGPSCoordinates": end_time_gps,
                    "isCopyPasted": is_likely_copy_paste,
                    "careGiverId": caregiver_id,
                    "typeOfService": type_of_service,
                }
            except Exception as exc:
                logger.warning("Error processing day column %d: %s", di, exc)
                return None

        # Run day columns in parallel (ThreadPoolExecutor mirrors Parallel.ForEach)
        day_results: list[Optional[dict]] = [None] * len(day_indices)
        with ThreadPoolExecutor(max_workers=self._max_workers) as executor:
            future_to_di = {executor.submit(process_day, di): di for di in day_indices}
            for future in as_completed(future_to_di):
                di = future_to_di[future]
                try:
                    day_results[di] = future.result()
                except Exception as exc:
                    logger.warning("Future error for day %d: %s", di, exc)

        # Order by DAY_MAP index (same as .NET OrderBy Array.IndexOf)
        for row in day_results:
            if row is not None:
                results.append(row)

        return results

    # ------------------------------------------------------------------
    # Buffer helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _extract_value_from_buffer(
        sheet_data: list[list], target: str, search_next_row: bool
    ) -> str:
        norm = target.strip().lower()
        for r_idx, row in enumerate(sheet_data):
            for c_idx, cell in enumerate(row):
                txt = _safe_str(cell).lower()
                if not txt or norm not in txt:
                    continue

                if search_next_row:
                    # Primary: cell below
                    if r_idx + 1 < len(sheet_data) and c_idx < len(sheet_data[r_idx + 1]):
                        v = _safe_str(sheet_data[r_idx + 1][c_idx])
                        if v:
                            return v
                    # Secondary: cell to the right
                    if c_idx + 1 < len(row):
                        v = _safe_str(row[c_idx + 1])
                        if v:
                            return v
                else:
                    # Primary: 5 cells to the right (matches .NET c+5)
                    if c_idx + 5 < len(row):
                        v = _safe_str(row[c_idx + 5])
                        if v:
                            return v
                    # Secondary: cell below
                    if r_idx + 1 < len(sheet_data) and c_idx < len(sheet_data[r_idx + 1]):
                        v = _safe_str(sheet_data[r_idx + 1][c_idx])
                        if v:
                            return v

                # Fallback: search up to 4 offsets
                for off in range(1, 5):
                    if c_idx + off < len(row):
                        v = _safe_str(row[c_idx + off])
                        if v:
                            return v
                    if r_idx + off < len(sheet_data) and c_idx < len(sheet_data[r_idx + off]):
                        v = _safe_str(sheet_data[r_idx + off][c_idx])
                        if v:
                            return v
        return ""

    @staticmethod
    def _extract_contract(sheet_data: list[list]) -> str:
        for row in sheet_data:
            for j, cell in enumerate(row):
                if _safe_str(cell).strip().lower() == "payer(s):":
                    # +4 first, then +1
                    if j + 4 < len(row):
                        v = _safe_str(row[j + 4])
                        if v:
                            return v
                    if j + 1 < len(row):
                        v = _safe_str(row[j + 1])
                        if v:
                            return v
        return "N/A"

    @staticmethod
    def _find_row_by_leading_label(sheet_data: list[list], label: str) -> int:
        for r_idx, row in enumerate(sheet_data):
            if not row:
                continue
            first = _safe_str(row[0]).strip()
            if first.lower().startswith(label.lower()):
                return r_idx
        return -1
