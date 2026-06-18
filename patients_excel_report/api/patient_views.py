"""
Port of PatientController.cs

Endpoint:
  POST /api/Patient/CheckSignatureFile
    - Accepts a .xlsx file or a .txt file containing the base64 of a .xlsx
    - Returns a JSON dict:  { "sheet1": { "1/5/2025": "signed", … }, … }

This is the simpler / legacy controller.  It only checks the Patient
Signature row and does NOT produce the full ReportRow audit output.
"""

import base64
import logging
import os
import tempfile
import uuid

from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView
from drf_spectacular.utils import extend_schema, OpenApiParameter
from drf_spectacular.types import OpenApiTypes

from patients_excel_report.services.image_extractor import (
    extract_images_anchored_to_worksheet,
)
from patients_excel_report.services.signature_detection_service import (
    SignatureDetectionService,
)

logger = logging.getLogger("patients_excel_report.patient_view")

# Day-column map  (Excel column letter → weekday key)
DAY_COLUMNS: list[tuple[str, str]] = [
    ("F", "Sun"),
    ("H", "Mon"),
    ("K", "Tue"),
    ("N", "Wed"),
    ("O", "Thu"),
    ("T", "Fri"),
    ("W", "Sat"),
]

_sig_service = SignatureDetectionService()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalize_whitespace(s: str) -> str:
    return " ".join(s.split()).strip() if s else ""


def _column_index_from_letter(col: str) -> int:
    """Convert column letter (A=1, B=2 …) to 1-based index."""
    result = 0
    for ch in col.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def _strip_data_prefix(s: str) -> str:
    if s.lower().startswith("data:"):
        idx = s.find(",")
        if idx >= 0:
            return s[idx + 1:]
    return s


def _find_patient_signature_row(ws) -> int:
    """
    Return the 1-based row number of the "Patient Signature" row,
    or 0 if not found.  Checks concatenated text of columns A–E.
    """
    for row in ws.iter_rows(max_col=5, values_only=True):
        combined = _normalize_whitespace(
            " ".join(str(c) if c is not None else "" for c in row)
        )
        if combined.lower() == "patient signature":
            return row[0]  # need the actual row number

    # openpyxl read_only iter_rows doesn't give row numbers; use cell iteration
    return 0


def _find_patient_signature_row_number(ws) -> int:
    from openpyxl import load_workbook

    last_row = ws.max_row or 0
    for r in range(1, last_row + 1):
        combined = _normalize_whitespace(
            " ".join(
                str(ws.cell(row=r, column=c).value or "")
                for c in range(1, 6)  # A=1 … E=5
            )
        )
        if combined.lower() == "patient signature":
            return r
    return 0

def _find_caregiver_signature_row_number(ws) -> int:
    last_row = ws.max_row or 0
    for r in range(1, last_row + 1):
        combined = _normalize_whitespace(
            " ".join(str(ws.cell(row=r, column=c).value or "") for c in range(1, 6))
        )
        if combined.lower() == "caregiver signature":
            return r
    return 0

def _process_excel_file(excel_path: str) -> dict:
    """
    Process all sheets and return a dict:
      { "sheet1": { "<date_label>": "signed"|"unsigned", … }, … }
    """
    from openpyxl import load_workbook

    output = {}

    wb = load_workbook(excel_path, data_only=True)

    for sheet_num, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]

        caregiver_row = _find_caregiver_signature_row_number(ws)
        sig_row_num = _find_patient_signature_row_number(ws)
        if sig_row_num == 0 and caregiver_row == 0:
            continue

        # Dates from row 15 for each day column
        date_map: dict[str, str] = {}
        for col_letter, _ in DAY_COLUMNS:
            col_idx = _column_index_from_letter(col_letter)
            cell_val = ws.cell(row=15, column=col_idx).value
            date_map[col_letter] = str(cell_val).strip() if cell_val else col_letter

        # Extract images for this sheet
        images = extract_images_anchored_to_worksheet(excel_path, sheet_name)

        caregiver_result: dict[str, str] = {}
        patient_result: dict[str, str] = {}
        for col_letter, day_key in DAY_COLUMNS:
            date_label = date_map.get(col_letter, day_key)

            if caregiver_row:
                img_bytes = images.get(f"{col_letter}{caregiver_row}")
                caregiver_result[date_label] = (
                    "signed" if img_bytes and _analyze_image(img_bytes) > 0.0 else "unsigned"
                )
            
            if sig_row_num:
                img_bytes = images.get(f"{col_letter}{sig_row_num}")
                patient_result[date_label] = (
                    "signed" if img_bytes and _analyze_image(img_bytes) > 0.0 else "unsigned"
                )
            cell_addr = f"{col_letter}{sig_row_num}"

        output[f"sheet{sheet_num}"] = {
            "caregiver_signature": caregiver_result,
            "patient_signature": patient_result,
        }

    wb.close()
    return output


def _analyze_image(image_bytes: bytes) -> float:
    """
    Returns the percentage of non-white/non-transparent pixels.
    Any value > 0 means the cell is considered signed.
    """
    from PIL import Image
    import io

    MAX_DIM = 300
    try:
        image = Image.open(io.BytesIO(image_bytes)).convert("RGBA")
        w, h = image.size

        if w > MAX_DIM or h > MAX_DIM:
            scale = min(MAX_DIM / w, MAX_DIM / h)
            image = image.resize((max(1, round(w * scale)), max(1, round(h * scale))), Image.LANCZOS)
            w, h = image.size

        total = w * h
        if total == 0:
            return 0.0

        pixels = image.load()
        non_empty = 0
        for y in range(h):
            for x in range(w):
                r, g, b, a = pixels[x, y]
                alpha = a / 255.0
                if alpha < 0.05:
                    continue
                lum = 0.2126 * (r / 255.0) + 0.7152 * (g / 255.0) + 0.0722 * (b / 255.0)
                if not (lum > 0.98 and alpha > 0.99):
                    non_empty += 1

        return non_empty * 100.0 / total
    except Exception:
        return 0.0


# ---------------------------------------------------------------------------
# View
# ---------------------------------------------------------------------------

class CheckSignatureFileView(APIView):
    """
    POST /api/Patient/CheckSignatureFile

    Upload a .xlsx file (or a .txt file containing the base64 of a .xlsx).
    Returns JSON: { "sheet1": { "<date>": "signed"|"unsigned" }, … }
    """

    parser_classes = [MultiPartParser]

    @extend_schema(
        summary="Check patient signatures (legacy, simple output)",
        description=(
            "Upload a .xlsx file or a .txt file containing the base64-encoded "
            "contents of a .xlsx.  Returns a dict mapping each sheet to a dict "
            "of date → 'signed'|'unsigned'."
        ),
        request={
            "multipart/form-data": {
                "type": "object",
                "properties": {
                    "file": {"type": "string", "format": "binary"},
                },
                "required": ["file"],
            }
        },
        responses={200: {"type": "object"}},
    )
    def post(self, request: Request) -> Response:
        file = request.FILES.get("file")

        if file is None:
            logger.debug("No file uploaded.")
            return Response(
                "No file uploaded. Upload a .txt file containing base64 of the "
                ".xlsx file.",
                status=status.HTTP_400_BAD_REQUEST,
            )

        fname = file.name or ""
        if not fname.lower().endswith(".txt") and not fname.lower().endswith(".xlsx"):
            logger.debug("Unsupported file type: %s", fname)
            return Response(
                "Please upload a .txt file containing base64 text or a .xlsx file.",
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            if fname.lower().endswith(".xlsx"):
                raw = file.read()
                b64 = base64.b64encode(raw).decode("ascii")
            else:
                b64 = file.read().decode("utf-8", errors="ignore")
        except Exception as exc:
            logger.error("Error reading uploaded file: %s", exc)
            return Response(
                f"Error reading uploaded file: {exc}",
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not b64 or not b64.strip():
            logger.warning("Uploaded file is empty.")
            return Response("Uploaded file is empty.", status=status.HTTP_400_BAD_REQUEST)

        b64 = _strip_data_prefix(b64).strip()
        return self._process_base64_and_respond(b64)

    def _process_base64_and_respond(self, b64: str) -> Response:
        logger.info("Processing base64 Excel data.")
        try:
            xlsx_bytes = base64.b64decode(b64)
        except Exception as exc:
            logger.error("Invalid base64 string: %s", exc)
            return Response(
                f"Provided string is not valid base64: {exc}",
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Verify ZIP magic bytes (PK header)
        if len(xlsx_bytes) < 4 or xlsx_bytes[0] != 0x50 or xlsx_bytes[1] != 0x4B:
            logger.warning("Provided data is not a valid .xlsx file.")
            return Response(
                "Provided data is not a valid .xlsx file (expected ZIP package).",
                status=status.HTTP_400_BAD_REQUEST,
            )

        temp_path = os.path.join(
            tempfile.gettempdir(), f"pschecker_{uuid.uuid4().hex}.xlsx"
        )
        try:
            with open(temp_path, "wb") as f:
                f.write(xlsx_bytes)

            result = _process_excel_file(temp_path)
            logger.info("Excel file processed successfully.")
            return Response(result, status=status.HTTP_200_OK)

        except Exception as exc:
            logger.error("Error processing file: %s", exc)
            return Response(
                f"Error processing file: {exc}",
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        finally:
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                    logger.info("Temp file deleted: %s", temp_path)
            except Exception as exc:
                logger.warning("Failed to delete temp file %s: %s", temp_path, exc)
